import openpyxl
from datetime import datetime, date, timedelta

def get_upcoming_events(days=14):
    """回傳未來N天的重要事件列表"""
    wb = openpyxl.load_workbook(r'C:\Users\User\.openclaw\workspace\2026_台股美股關鍵事件曆_v2.xlsx')
    ws = wb.active
    today = date.today()
    events = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] is None: continue
        try:
            d_str = str(row[0]).strip()
            if len(d_str) != 10 or d_str[4] != '-' or d_str[7] != '-': continue
            d = datetime.strptime(d_str, '%Y-%m-%d').date()
            diff = (d - today).days
            if 0 <= diff <= days:
                events.append({
                    'date': d,
                    'diff': diff,
                    'name': str(row[1]) if row[1] else '',
                    'category': str(row[2]) if row[2] else '',
                    'impact': str(row[3]) if row[3] else '',
                    'detail': str(row[4])[:60] if row[4] else '',
                })
        except:
            pass
    events.sort(key=lambda x: x['diff'])
    return events

if __name__ == '__main__':
    evts = get_upcoming_events()
    print(f'未來兩週重要事件 ({len(evts)}筆):')
    for e in evts:
        print(f'  +{e["diff"]}d {e["date"]} | {e["impact"]} | {e["name"]}')
