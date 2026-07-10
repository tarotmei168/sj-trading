"""
Excel版晨報 — 手機可直接開啟
每天8:30自動產出，不花token
"""
import os, json
from datetime import datetime, timedelta
from dotenv import load_dotenv
import shioaji as sj
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()
BASE = r"C:\Users\User\.openclaw\workspace\sj-trading"
now = datetime.now()
OUTPUT_FILE = os.path.join(BASE, f"{now.strftime('%m%d')}_晨報.xlsx")

def login():
    api = sj.Shioaji(simulation=True)
    api.login(api_key=os.environ['SJ_API_KEY'], secret_key=os.environ['SJ_SEC_KEY'])
    return api

def get_data(api, sid):
    contract = api.Contracts.Stocks[sid]
    end = datetime.now()
    start = end - timedelta(days=15)
    all_dfs = []
    seg_end = end
    while seg_end > start:
        seg_start = max(seg_end - timedelta(days=14), start)
        try:
            kbars = api.kbars(contract=contract, start=seg_start.strftime("%Y-%m-%d"), end=seg_end.strftime("%Y-%m-%d"))
            if len(kbars.ts)==0: break
            df = pd.DataFrame({"datetime":pd.to_datetime(kbars.ts),"open":kbars.Open,"high":kbars.High,"low":kbars.Low,"close":kbars.Close,"volume":kbars.Volume})
            all_dfs.append(df)
            seg_end = seg_start - timedelta(seconds=1)
        except: break
    if not all_dfs: return None
    min_df = pd.concat(all_dfs)
    min_df.drop_duplicates(subset=["datetime"],inplace=True)
    min_df.sort_values("datetime",inplace=True)
    min_df.set_index("datetime",inplace=True)
    df30 = min_df.resample("30min").agg({"open":"first","high":"max","low":"min","close":"last","volume":"sum"}).dropna()
    if len(df30)<15: return None
    low_min=df30["low"].rolling(9).min(); high_max=df30["high"].rolling(9).max()
    rsv=((df30["close"]-low_min)/(high_max-low_min))*100; rsv=rsv.fillna(50)
    k_vals,d_vals=[50]*9,[50]*9
    for i in range(9,len(df30)):
        k_new=(2/3)*k_vals[-1]+(1/3)*rsv.iloc[i]; d_new=(2/3)*d_vals[-1]+(1/3)*k_new
        k_vals.append(k_new);d_vals.append(d_new)
    df30["K"]=k_vals;df30["D"]=d_vals
    df30["vol_ma5"]=df30["volume"].rolling(5).mean()
    df30["vol_ratio"]=df30["volume"]/df30["vol_ma5"].replace(0,np.nan)
    return df30

def analyze(sid, df):
    if df is None or len(df)<12: return None
    last=df.iloc[-1]; prev=df.iloc[-2]
    k=round(last["K"],1); d=round(last["D"],1)
    kp=round(prev["K"],1); dp=round(prev["D"],1)
    p=round(last["close"],2)
    vr=round(last["vol_ratio"],2) if not pd.isna(last.get("vol_ratio",np.nan)) else 0
    sig=""
    if kp<=dp and k>d: sig="黃金交叉"
    elif kp>=dp and k<d: sig="死亡交叉"
    else: sig="盤整"
    ks=""
    if k>80:ks="高檔⚠️"
    elif k>60:ks="偏高"
    elif k>40:ks="中位"
    elif k>20:ks="偏低"
    else:ks="超賣💡"
    vs=""
    if vr>1.5:vs="爆量🔥"
    elif vr>0.8:vs="正常"
    else:vs="量縮"
    return {"price":p,"k":k,"d":d,"k_pos":ks,"vol_r":vr,"vol_s":vs,"signal":sig}

# 股票清單 + 股性備註
STOCKS = [
    ("3711","日月光","封測龍頭","K81超買，高檔注意回檔"),
    ("4958","臻鼎KY","PCB大廠","K<D偏空，量縮防守570"),
    ("3042","晶技","石英元件","K值急升，量價背離要注意"),
    ("2337","旺宏","記憶體","超賣區K=17，等金叉反彈"),
    ("2436","偉詮電","IC設計","高檔盤整，守MA20續抱"),
    ("3673","TPKKY","PCB小而美","買氣93%超強，但高檔死叉"),
    ("5351","鈺創","利基IC設計","K=93極高檔，量縮防殺多"),
    ("2317","鴻海","AI權值","超賣區K=15，不殺低"),
    ("2454","聯發科","IC設計龍頭","短線死叉，日K多頭不變"),
    ("8150","南茂","封測","區間90~120，等95以下"),
    ("2330","台積電","權值龍頭","爆量2.6倍，中位盤整"),
    ("4961","天鈺","驅動IC","爆量1.89倍，守165噴出"),
    ("6451","訊芯KY","CPO概念","等回測500~510再進"),
    ("3532","台勝科","矽晶圓","黃金交叉，底部成形"),
    ("2327","國巨","被動元件","K=16超賣，等金叉反彈"),
    ("8016","矽創","IC設計","量能回升，大戶吸籌"),
    ("2464","盟立","機器人","K=13極超賣，左側埋伏"),
    ("6139","亞翔","建廠設備","高檔死叉，避開"),
    ("1303","南亞","塑化","高檔死叉量縮，避開"),
    ("5425","台半","二極體","K=96量縮假突破，避開"),
]

print("📊 正在產出Excel晨報...", flush=True)
api = login()

rows_data = []
for sid, name, nature, strategy in STOCKS:
    df = get_data(api, sid)
    info = analyze(sid, df)
    if info:
        rows_data.append([sid, name, nature, info["price"], info["k"], info["d"], info["k_pos"], info["vol_r"], info["vol_s"], info["signal"], strategy])
        print(f"  {name}({sid}) K={info['k']} {info['signal']}", flush=True)

api.logout()

# 建立 Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"{now.strftime('%m%d')}晨報"

# 樣式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A237E", end_color="283593", fill_type="solid")
green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
thin_border = Border(left=Side(style='thin',color='00DDDDDD'), right=Side(style='thin',color='00DDDDDD'), top=Side(style='thin',color='00DDDDDD'), bottom=Side(style='thin',color='00DDDDDD'))

# 標題
ws.merge_cells('A1:K1')
title_cell = ws['A1']
title_cell.value = f"📊 晨報系統 {now.strftime('%Y/%m/%d')} — 30分K即時分析 (資料來源：永豐金Shioaji API)"
title_cell.font = Font(bold=True, size=14, color="1A237E")
title_cell.alignment = Alignment(horizontal='center')

# 表頭
headers = ["代號","名稱","股性","股價","K值","D值","K位置","量比","量能","30分K訊號","操作策略"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# 資料
for i, row_data in enumerate(rows_data, 4):
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        # 顏色標記
        if row_data[9] == "黃金交叉":
            cell.fill = green_fill
        elif row_data[9] == "死亡交叉":
            cell.fill = red_fill
        # 量能顏色
        if col == 9:
            if val == "爆量🔥": cell.font = Font(color="2E7D32", bold=True)
            elif val == "量縮": cell.font = Font(color="C62828")
    # 代號+名稱粗體
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=2).font = Font(bold=True)

# 欄寬
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 16
ws.column_dimensions['K'].width = 35

# 註解
ws.cell(row=len(rows_data)+5, column=1, value="📌 說明：").font = Font(bold=True, size=10)
notes = [
    "🟢 黃金交叉+量增 = 最佳買點",
    "🔴 死亡交叉+放量 = 建議出場",
    "🟡 黃金交叉+量縮 = 假突破嫌疑",
    "💡 超賣區(K<20) = 留意反彈",
    "⚠️ 高檔區(K>80) = 注意回檔",
    "🔥 爆量(量比>1.5) = 方向明確",
]
for i, note in enumerate(notes):
    ws.cell(row=len(rows_data)+6+i, column=1, value=note)

wb.save(OUTPUT_FILE)
print(f"\n✅ {OUTPUT_FILE}")
