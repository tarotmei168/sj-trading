# -*- coding: utf-8 -*-
"""HTML晨報 v8 — 完整版：股表+新聞+夜盤+事件"""
import os, re, json, urllib.request
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
import shioaji as sj
import pandas as pd
import numpy as np
import openpyxl

load_dotenv()
BASE=r"C:\Users\User\.openclaw\workspace\sj-trading"
now=datetime.now()
OUTPUT_HTML=os.path.join(BASE,f"{now.strftime('%m%d')}_晨報.html")
OUTPUT_XLSX=os.path.join(BASE,f"{now.strftime('%m%d')}_晨報.xlsx")

# ===== 爬蟲 =====
def get_news(limit=6):
    url="https://news.cnyes.com/api/v3/news/category/tw_stock?limit=20&page=1"
    try:
        req=urllib.request.Request(url,headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req,timeout=5) as r:
            data=json.loads(r.read().decode("utf-8"))
        items=data.get("items",{}).get("list",[]) if "items" in data else data.get("list",[])
        return [{"title":item.get("title",""),"date":str(item.get("publishedAt",""))[:10]} for item in items[:limit]]
    except: return [{"title":"連線失敗","date":""}]

def get_sox():
    try:
        req=urllib.request.Request("https://www.google.com/finance/quote/SOXX:INDEXNASDAQ",headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req,timeout=5) as r:
            h=r.read().decode("utf-8","ignore")
        m=re.search(r'class="YMlKec"[^>]*>([\d,.]+)',h)
        return m.group(1).replace(",","") if m else "查無"
    except: return "連線失敗"

def get_events():
    """從事件曆抓未來兩週事件"""
    try:
        wb=openpyxl.load_workbook(r"C:\Users\User\.openclaw\workspace\2026_台股美股關鍵事件曆_v2.xlsx")
        ws=wb.active
        today=date.today()
        events=[]
        for row in ws.iter_rows(min_row=1,max_row=ws.max_row,values_only=True):
            if row[0] is None: continue
            d_str=str(row[0]).strip()
            if len(d_str)!=10 or d_str[4]!='-' or d_str[7]!='-': continue
            d=datetime.strptime(d_str,"%Y-%m-%d").date()
            diff=(d-today).days
            if 0<=diff<=14:
                events.append({"date":d.strftime("%m/%d"),"diff":diff,"name":str(row[1]) if row[1] else "","impact":str(row[3]) if row[3] else "","category":str(row[2]) if row[2] else ""})
        events.sort(key=lambda x:x["diff"])
        return events
    except: return []

# ===== Shioaji =====
def login():
    api=sj.Shioaji(simulation=True)
    api.login(api_key=os.environ['SJ_API_KEY'],secret_key=os.environ['SJ_SEC_KEY'])
    return api

def get30(api,sid,days=15):
    contract=api.Contracts.Stocks[sid];end=datetime.now();start=end-timedelta(days=days)
    ads,se=[],end
    while se>start:
        ss=max(se-timedelta(days=14),start)
        try:
            k=api.kbars(contract=contract,start=ss.strftime("%Y-%m-%d"),end=se.strftime("%Y-%m-%d"))
            if len(k.ts)==0:break
            ads.append(pd.DataFrame({"ts":pd.to_datetime(k.ts),"o":k.Open,"h":k.High,"l":k.Low,"c":k.Close,"v":k.Volume}))
            se=ss-timedelta(seconds=1)
        except:break
    if not ads: return None
    m=pd.concat(ads);m.drop_duplicates(subset=["ts"],inplace=True);m.sort_values("ts",inplace=True);m.set_index("ts",inplace=True)
    d30=m.resample("30min").agg({"o":"first","h":"max","l":"min","c":"last","v":"sum"}).dropna()
    if len(d30)<15: return None
    lm=d30["l"].rolling(9).min();hm=d30["h"].rolling(9).max()
    rsv=((d30["c"]-lm)/(hm-lm))*100;rsv=rsv.fillna(50)
    kv,dv=[50]*9,[50]*9
    for i in range(9,len(d30)):
        kn=(2/3)*kv[-1]+(1/3)*rsv.iloc[i];dn=(2/3)*dv[-1]+(1/3)*kn
        kv.append(kn);dv.append(dn)
    d30["K"]=kv;d30["D"]=dv;d30["KDgap"]=np.array(kv)-np.array(dv)
    e12=d30["c"].ewm(span=12).mean();e26=d30["c"].ewm(span=26).mean()
    d30["MACD"]=e12-e26;d30["MACDs"]=d30["MACD"].ewm(span=9).mean();d30["MACDh"]=d30["MACD"]-d30["MACDs"]
    dl=d30["c"].diff();g=dl.clip(lower=0);ls=-dl.clip(upper=0)
    d30["RSI"]=100-(100/(1+g.rolling(14).mean()/ls.rolling(14).mean().replace(0,np.nan)))
    d30["vm5"]=d30["v"].rolling(5).mean();d30["vr"]=d30["v"]/d30["vm5"].replace(0,np.nan)
    return d30

def calc(sid,d30):
    if d30 is None or len(d30)<14: return None
    l=d30.iloc[-1];p=d30.iloc[-2]
    k=round(l["K"],1);d=round(l["D"],1);kp=round(p["K"],1);dp=round(p["D"],1)
    pr=round(l["c"],2);vr=round(l["vr"],2) if not pd.isna(l.get("vr",np.nan)) else 0
    gap=round(l["KDgap"],1);gp=round(p["KDgap"],1)
    mh=round(l["MACDh"],2);mhp=round(p["MACDh"],2)
    rsi=round(l["RSI"],1) if not pd.isna(l.get("RSI",np.nan)) else 50
    rsi_p=round(p["RSI"],1)
    if kp<=dp and k>d: kd="🟢金叉" if vr>1.0 else "🟡金叉量縮"
    elif kp>=dp and k<d: kd="🔴死叉"
    elif gap<0 and gap>gp and gap>-3 and k<50: kd="💡即將金叉"
    else: kd="⚪盤整"
    if k>80: kd+="(高)"
    elif k<20: kd+="(低)"
    if rsi>70: rs=f"RSI{rsi}🔥超買"
    elif rsi<30: rs=f"RSI{rsi}💧超賣"
    elif rsi>50: rs=f"RSI{rsi}📈偏多"
    else: rs=f"RSI{rsi}📉偏空"
    macdv=round(l["MACD"],2)
    if macdv>0: ms="🔴紅柱↑" if mh>mhp else "🔴紅柱↓"
    else: ms="🟢綠柱↑" if mh<mhp else "🟢綠柱↓"
    if mh<0 and macdv>0: ms="🔄翻紅"
    if mh>0 and macdv<0: ms="🔄翻綠"
    vs="🔥爆量" if vr>1.5 else ("✅正常" if vr>0.8 else "💤量縮")
    hints=[]
    if "金叉" in kd and "量縮" not in kd: hints.append("🟢金叉+量增=強力買點")
    elif "金叉量縮" in kd: hints.append("🟡金叉量縮=假突破")
    elif "死叉" in kd and vr>1.0: hints.append("🔴死叉放量=出場")
    elif "死叉" in kd: hints.append("🟠死叉觀察")
    if "即將金叉" in kd: hints.append("💡準備埋伏資金")
    if rsi>70 and rsi<rsi_p: hints.append("⚠️RSI背離")
    if rsi<30 and rsi>rsi_p: hints.append("💡RSI背離增")
    if k>80 and vr<0.8: hints.append("⚠️高檔量縮")
    if k<20 and vr>1.0: hints.append("💡超賣放量")
    if not hints: hints.append("⚪觀望")
    return {"price":pr,"kd":kd,"rs":rs,"ms":ms,"vs":vs,"hint":" | ".join(hints[:3])}

NAMES={"3711":"日月光","4958":"臻鼎KY","3042":"晶技","2337":"旺宏","2436":"偉詮電","3673":"TPKKY","5351":"鈺創","2317":"鴻海","2454":"聯發科","8150":"南茂","2330":"台積電","4961":"天鈺","6451":"訊芯KY","3532":"台勝科","2327":"國巨","8016":"矽創","2464":"盟立","6139":"亞翔","1303":"南亞","5425":"台半導"}
TYPES={"3711":"權值","4958":"穩健","3042":"績優","2337":"循環","2436":"題材","3673":"活潑","5351":"主力","2317":"權值","2454":"龍頭","8150":"循環","2330":"龍頭","4961":"活潑","6451":"波動","3532":"循環","2327":"循環","8016":"績優","2464":"波動","6139":"波動","1303":"循環","5425":"循環"}
STRATS={"3711":"守680跌破減碼","4958":"防守570等金叉","3042":"守紅K低點","2337":"等量增金叉","2436":"守MA20續抱","3673":"防守75元","5351":"高檔防殺多","2317":"不殺低等金叉","2454":"日K續抱","8150":"等95以下金叉","2330":"低點不破可進","4961":"守165噴出","6451":"等500~510止跌","3532":"底部成形可試","2327":"等金叉反彈","8016":"大戶吸籌觀察","2464":"左側埋伏","6139":"勿追高","1303":"波段見頂避開","5425":"明日砸盤避開"}

def make_table(data,title,icon):
    h=f'<div class="section"><h2>{icon} {title}</h2><div class="table-wrap"><table><thead><tr>'
    for c in ["代號","名稱","股性","股價","30分KD","RSI","MACD","量能","🎯提示","策略"]:
        h+=f'<th>{c}</th>'
    h+='</tr></thead><tbody>'
    for r in data:
        sid,nm,ty=r[0],NAMES.get(r[0],r[0]),TYPES.get(r[0],"")
        st=STRATS.get(r[0],"")
        cls="death" if "🔴" in str(r[2]) else ("golden" if "🟢" in str(r[2]) or "💡" in str(r[2]) else "")
        h+=f'<tr class="{cls}"><td class="sid">{sid}</td><td class="name">{nm}</td><td class="type">{ty}</td>'
        h+=f'<td class="price">{r[1]}</td>'
        for c in range(2,7): h+=f'<td>{r[c]}</td>'
        h+=f'<td class="strategy">{st}</td></tr>'
    h+='</tbody></table></div></div>'
    return h

print("📊 產出完整晨報 v8...")
api=login()

CORE=[];WATCH=[]
for sid in ["3711","4958","3042","2337","2436","3673","5351","2317","2454","8150"]:
    d30=get30(api,sid);info=calc(sid,d30)
    if info: CORE.append([sid]+[info["price"],info["kd"],info["rs"],info["ms"],info["vs"],info["hint"]])
for sid in ["2330","4961","6451","3532","2327","8016","2464","6139","1303","5425"]:
    d30=get30(api,sid);info=calc(sid,d30)
    if info: WATCH.append([sid]+[info["price"],info["kd"],info["rs"],info["ms"],info["vs"],info["hint"]])
api.logout()

# 爬蟲資料
sox=get_sox()
news=get_news()
events=get_events()

# 目前熱門題材（每日更新）
HOT_TOPICS = [
    "🔥 被動元件：國巨/信昌電 缺貨漲價題材持續",
    "🔥 功率元件：台半/強茂 車用功率半導體需求旺",
    "🔥 成熟製程：聯電/力積電 產能利用率回升",
    "🔥 AI伺服器：鴻海/廣達 輝達新晶片帶動",
]

html=f'''<!DOCTYPE html><html lang="zh-TW"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=3.0">
<title>晨報 {now.strftime('%m%d')}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,"Microsoft JhengHei",sans-serif;background:#f0f2f5;padding:12px;color:#333;font-size:18px}}
.container{{max-width:100%;margin:0auto}}
.header{{background:linear-gradient(135deg,#1a237e,#283593);color:#fff;padding:24px;border-radius:14px;margin-bottom:16px}}
.header h1{{font-size:26px;margin-bottom:6px}}
.header p{{font-size:16px;opacity:.85}}
.section{{background:#fff;border-radius:14px;padding:16px;margin-bottom:16px;box-shadow:0 2px 8px rgba(0,0,0,.1)}}
.section h2{{font-size:20px;margin-bottom:12px;padding-bottom:8px;border-bottom:2px solid #e0e0e0}}
.table-wrap{{overflow-x:auto;width:100%}}
table{{width:100%;border-collapse:collapse;font-size:17px;min-width:100%}}
th{{background:#1a237e;color:#fff;padding:12px 8px;text-align:left;font-weight:600;font-size:16px;position:sticky;top:0;white-space:nowrap}}
td{{padding:12px 8px;border-bottom:1px solid #e0e0e0;vertical-align:middle;white-space:nowrap}}
tr:hover{{background:#e8eaf6}}
.golden{{background:#e8f5e9!important}}
.death{{background:#ffebee!important}}
.sid{{font-weight:700;font-size:16px}}
.name{{font-weight:700;font-size:17px}}
.type{{font-size:15px;color:#666}}
.price{{font-size:18px;font-weight:700}}
.strategy{{font-size:16px;color:#1565c0;line-height:1.5;white-space:normal;min-width:120px}}
.footer{{text-align:center;padding:20px;color:#999;font-size:14px}}
.news-item{{padding:6px 0;font-size:16px;border-bottom:1px solid #f0f0f0}}
.event-date{{font-weight:700;color:#1a237e;min-width:50px}}
.event-impact{{min-width:80px}}
</style></head><body><div class="container">

<div class="header"><h1>📊 晨報 {now.strftime('%m/%d')}</h1>
<p>KD+RSI+MACD+量能 ｜ 永豐金API即時 | 費半:{sox}</p></div>
'''

# 夜盤+費半+題材
html+=f'<div class="section"><h2>🌙 隔夜盤勢 & 熱門題材</h2>'
html+=f'<table><tbody>'
for t in HOT_TOPICS:
    html+=f'<tr><td style="font-size:17px;padding:6px">{t}</td></tr>'
html+=f'</tbody></table></div>'

# 股票表
html+=make_table(CORE,"💼 核心持股追蹤","🏆")
html+=make_table(WATCH,"🔭 潛力波段標的","🚀")

# 新聞
html+=f'<div class="section"><h2>📰 即時新聞</h2>'
for n in news:
    html+=f'<div class="news-item">📰 [{n["date"]}] {n["title"]}</div>'
html+=f'</div>'

# 事件
html+=f'<div class="section"><h2>📅 未來兩週重要事件</h2><table><thead><tr><th>日期</th><th>事件</th><th>分類</th><th>影響</th></tr></thead><tbody>'
for e in events:
    bg=""
    if "🔥" in e["impact"]: bg='style="background:#fff3e0"'
    elif "🔴" in e["impact"]: bg='style="background:#ffebee"'
    elif "🟢" in e["impact"]: bg='style="background:#e8f5e9"'
    html+=f'<tr {bg}><td class="event-date"><b>{e["date"]}</b></td><td>{e["name"]}</td><td>{e["category"]}</td><td class="event-impact">{e["impact"]}</td></tr>'
html+=f'</tbody></table></div>'

# 心法
html+='<div class="section"><h2>📌 操盤心法</h2><table><tbody>'
for t in [
    "🟢 金叉+量增+MACD紅柱 = 強力買點",
    "🟡 金叉量縮+MACD綠柱 = 假突破90%",
    "🔴 死叉+放量+MACD翻綠 = 強烈賣出",
    "💡 KD即將金叉+量微增 = 提前埋伏",
    "🔥 被動元件/功率元件/成熟製程為當前主流",
]:
    html+=f'<tr><td style="font-size:17px;padding:8px">{t}</td></tr>'
html+='</tbody></table></div>'

html+=f'<div class="footer">晨報 v8 ｜ {now.strftime("%Y/%m/%d %H:%M")} ｜ 永豐金Shioaji API即時</div>'
html+='</div></body></html>'

with open(OUTPUT_HTML,"w",encoding="utf-8") as f: f.write(html)
print(f"✅ {OUTPUT_HTML}")
