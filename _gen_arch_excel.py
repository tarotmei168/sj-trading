#!/usr/bin/env python3
"""把 architecture_master.md 轉成 Excel，方便主人修改"""
import os, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = r'C:\Users\User\.openclaw\workspace\sj-trading'
MD_PATH = os.path.join(BASE_DIR, 'architecture_master.md')
XLSX_PATH = os.path.join(BASE_DIR, 'output', '架構表.xlsx')

wb = openpyxl.Workbook()

# ── 樣式 ──
hdr_font = Font(bold=True, size=12, color='FFFFFF')
hdr_fill = PatternFill(start_color='CC3333', end_color='CC3333', fill_type='solid')
cat_font = Font(bold=True, size=12, color='CC3333')
cat_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_cat(ws, row, col, val):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = cat_font
    cell.fill = cat_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')

def add_row(ws, row, data, bold_first=True):
    for i, v in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if bold_first and i == 1:
            cell.font = Font(bold=True)

# ═══════════════════════════════════════════
# Sheet 1: 時程排程
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = '時程排程'
ws1.append(['時間', '任務', '指令'])
style_header(ws1, 1, 3)
rows = [
    ['08:30 🦞', '早報產出 + Git Push', 'daily_web_report.py'],
    ['08:30~13:30 📡', '盤中30分K KD監控（每5分鐘）', 'cron 啟動'],
    ['16:30 📊', '全市場投信掃描 + 更新早報 + git push', 'daily_market_update.py → daily_web_report.py'],
]
for i, r in enumerate(rows, 2):
    add_row(ws1, i, r)

# ═══════════════════════════════════════════
# Sheet 2: 總經風控閾值
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('總經風控')
ws2.append(['指標', '閾值', '行動'])
style_header(ws2, 1, 3)
rows = [
    ['費半SOX (唯一關注指數)', '>±3%', '開盤基調強烈標記'],
    ['台指期', '盤前08:45前後', '開盤基調判斷'],
]
for i, r in enumerate(rows, 2):
    add_row(ws2, i, r)

# ═══════════════════════════════════════════
# Sheet 3: 核心持股 + 潛力股
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('核心持股')
ws3.append(['類型', '代號', '名稱', '產業', '最佳K值', 'VolF', '位置過濾', '3年勝率', '3年總報酬'])
style_header(ws3, 1, 8)
rows = [
    ['核心', '2436', '偉詮電', 'PC週邊IC', 5, '無', '不限', '35.6%', '+91.1%'],
    ['核心', '2337', '旺宏', '記憶體', 21, '無', '不限', '34.0%', '+282.9%'],
    ['核心', '5351', '鈺創', '記憶體', 14, '無', '不限', '36.7%', '+584.7%'],
    ['核心', '3673', 'TPK-KY', '觸控', 14, '1.5x', '不限', '39.0%', '+47.7%'],
    ['核心', '3711', '日月光', '封測', 21, '無', '不限', '31.7%', '+30.5%'],
    ['核心', '4958', '臻鼎-KY', 'PCB', 21, '無', 'K<50', '31.1%', '+91.5%'],
    ['核心', '3042', '晶技', '石英元件', 14, '1.5x', '不限', '34.0%', '+25.9%'],
    ['核心', '2454', '聯發科', 'IC設計', 21, '無', '不限', '30.3%', '+57.4%'],
    ['核心', '2317', '鴻海', '電子代工', 14, '1.5x', '不限', '33.3%', '+19.3%'],
    ['核心', '8150', '南茂', '封測', 21, '無', '不限', '34.6%', '+108.7%'],
    ['核心', '2330', '台積電', '晶圓代工', 9, '1.5x', '不限', '41.4%', '+32.7%'],
]
for i, r in enumerate(rows, 2):
    add_row(ws3, i, r)

# 潛力股備註
ws3.cell(row=len(rows)+3, column=1, value='🎯 潛力股規則').font = cat_font
ws3.cell(row=len(rows)+4, column=1, value='每天 16:30 從 TWSE T86 全市場掃描')
ws3.cell(row=len(rows)+5, column=1, value='篩選：非核心持股、投信連買≥3天、總額>50萬、前10名')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 14

# ═══════════════════════════════════════════
# Sheet 4: KD 策略
# ═══════════════════════════════════════════
ws4 = wb.create_sheet('KD策略')
ws4.append(['條件', '策略'])
style_header(ws4, 1, 2)
rows = [
    ['K金叉(gap>3) + RSI<60', '🟢 K金叉 可持股'],
    ['逼近金叉(gap>0) + RSI<50', '🟡 近金叉 觀望期待'],
    ['死叉(gap<-3)', '🔴 死叉中 避開'],
    ['RSI>70', '🔴 RSI過熱 注意回檔'],
    ['RSI<30 + 金叉', '🟢 RSI超賣+金叉 留意買點'],
    ['其他', '➖ 觀望'],
]
for i, r in enumerate(rows, 2):
    add_row(ws4, i, r)
ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 30

# ═══════════════════════════════════════════
# Sheet 5: 早報版型規範
# ═══════════════════════════════════════════
ws5 = wb.create_sheet('早報版型')
ws5.append(['順序', '區塊', '內容'])
style_header(ws5, 1, 3)
rows = [
    ['1', '費半SOX + 台指期', '開盤基調'],
    ['2', '核心持股表格', '代號/名稱/股價/KD(3年最佳K值)/RSI/量能/提示/策略'],
    ['3', '未來14天事件', '除息/法說/FOMC/季底'],
    ['4', '台美產業聯動', '美股波動地圖'],
    ['5', '投信法人建倉', '投信買超 + 法人(外資) 各10檔'],
    ['6', '潛力股候選', '投信連買非持股，含KD/RSI'],
    ['7', '川普投顧', '關稅/晶片禁令'],
]
for i, r in enumerate(rows, 2):
    add_row(ws5, i, r)
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 22
ws5.column_dimensions['C'].width = 45

# ═══════════════════════════════════════════
# Sheet 6: 數據源架構
# ═══════════════════════════════════════════
ws6 = wb.create_sheet('數據源')
ws6.cell(row=1, column=1, value='🎯 Shioaji（即時報價 + 歷史30分K KD）').font = cat_font
ws6.append(['功能', '說明'])
style_header(ws6, 2, 2)
rows = [
    ['即時快照 api.snapshots()', '11檔核心持股現價/漲跌/最高最低/成交量'],
    ['歷史1分K api.kbars()', '每次14天，分79段拉取3年，合成30分K KD'],
    ['開盤基調 check_market_tone()', '用台積電即時漲跌判斷'],
    ['database/3y_kd/', '每檔約6,500根30分K KD（3年回測用）'],
]
for i, r in enumerate(rows, 3):
    add_row(ws6, i, r)

r = len(rows) + 4
ws6.cell(row=r, column=1, value='🎯 技術指標（完全本機離線計算）').font = cat_font
ws6.cell(row=r+1, column=1, value='指標')
ws6.cell(row=r+1, column=2, value='計算方式')
ws6.cell(row=r+1, column=3, value='資料源')
style_header(ws6, r+1, 3)
rows2 = [
    ['KD', '30分K KD，每檔使用3年回測最佳K值', 'database/3y_kd/{sid}_kd.csv'],
    ['RSI', '日收盤價14期RSI', 'FinMind API（60天日K）'],
    ['量能', '當前量/前5根均量，<0.8=量縮 >1.5=放量', '30分K volume'],
]
for i, r2 in enumerate(rows2, r+2):
    add_row(ws6, i, r2)

ws6.column_dimensions['A'].width = 35
ws6.column_dimensions['B'].width = 45
ws6.column_dimensions['C'].width = 30

# ═══════════════════════════════════════════
# Sheet 7: 檔案組織
# ═══════════════════════════════════════════
ws7 = wb.create_sheet('檔案組織')
ws7.append(['路徑', '說明'])
style_header(ws7, 1, 2)
rows = [
    ['MEMORY.md', '主人偏好/工作模式/通訊頻道'],
    ['MORNING_CHECKLIST.md', '開機記憶卡（完整設定）'],
    ['HEARTBEAT.md', '復活指令'],
    ['architecture_master.md', '本文件（最高準則）'],
    ['web/index.html', '早報（GitHub Pages）'],
    ['database/3y_kd/', '11檔×6500根30分K KD'],
    ['output/trust_scan_latest.json', '投信掃描結果'],
    ['output/SITC_Accumulation.csv', '投信買賣超累積'],
    ['output/bt_3y_kd_report.html', '3年回測報告'],
    ['src/sj_trading/daily_web_report.py', '早報產生器（主程式）'],
    ['src/sj_trading/daily_market_update.py', '16:30 投信掃描'],
    ['src/sj_trading/global_weather.py', '總經氣象台（SOX+台指期+事件）'],
    ['src/sj_trading/morning_news.py', '新聞引擎（鉅亨網過濾中國）'],
    ['src/sj_trading/calc_trust_rate.py', '股本滲透率'],
    ['src/sj_trading/shioaji_helper.py', '永豐金API'],
    ['src/sj_trading/us_tw_mapping_matrix.py', '台美聯動40組'],
    ['src/sj_trading/day_engine_v2.py', '盤中監控引擎'],
    ['src/sj_trading/download_3y_intraday_kd.py', '3年KD下載'],
]
for i, r in enumerate(rows, 2):
    add_row(ws7, i, r)
ws7.column_dimensions['A'].width = 45
ws7.column_dimensions['B'].width = 40

# ── 存檔 ──
wb.save(XLSX_PATH)
print(f'OK 已產出: {XLSX_PATH}')
print(f'   共 {len(wb.sheetnames)} 個 Sheet: {", ".join(wb.sheetnames)}')
