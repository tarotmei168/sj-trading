# -*- coding: utf-8 -*-
"""
Excel晨報 v6
"""
import os
from datetime import datetime, timedelta
from dotenv import load_dotenv
import shioaji as sj
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()
BASE = r"C:\Users\User\.openclaw\workspace\sj-trading"
now = datetime.now()
OUTPUT = os.path.join(BASE, f"{now.strftime('%m%d')}_晨報.xlsx")

def login():
    api = sj.Shioaji(simulation=True)
    api.login(api_key=os.environ['SJ_API_KEY'], secret_key=os.environ['SJ_SEC_KEY'])
    return api

def get30(api, sid, days=15):
    contract=api.Contracts.Stocks[sid]; end=datetime.now(); start=end-timedelta(days=days)
    ads,se=[],end
    while se>start:
        ss=max(se-timedelta(days=14),start)
        try:
            k=api.kbars(contract=contract,start=ss.strftime("%Y-%m-%d"),end=se.strftime("%Y-%m-%d"))
            if len(k.ts)==0: break
            ads.append(pd.DataFrame({"ts":pd.to_datetime(k.ts),"o":k.Open,"h":k.High,"l":k.Low,"c":k.Close,"v":k.Volume}))
            se=ss-timedelta(seconds=1)
        except: break
    if not ads: return None
    m=pd.concat(ads);m.drop_duplicates(subset=["ts"],inplace=True);m.sort_values("ts",inplace=True);m.set_index("ts",inplace=True)
    d30=m.resample("30min").agg({"o":"first","h":"max","l":"min","c":"last","v":"sum"}).dropna()
    if len(d30)<15: return None
    lm=d30["l"].rolling(9).min();hm=d30["h"].rolling(9).max()
    rsv=((d30["c"]-lm)/(hm-lm))*100;rsv=rsv.fillna(50)
    kv,dv=[50]*9,[50]*9
    for i in range(9,len(d30)):
        kn=(2/3)*kv[-1]+(1/3)*rsv.iloc[i];dn=(2/3)*dv[-1]+(1/3)*kn
        kv.append(kn);dv.append(dn)
    d30["K"]=kv;d30["D"]=dv;d30["KDgap"]=np.array(kv)-np.array(dv)
    e12=d30["c"].ewm(span=12).mean();e26=d30["c"].ewm(span=26).mean()
    d30["MACD"]=e12-e26;d30["MACDs"]=d30["MACD"].ewm(span=9).mean();d30["MACDh"]=d30["MACD"]-d30["MACDs"]
    dl=d30["c"].diff();g=dl.clip(lower=0);ls=-dl.clip(upper=0)
    d30["RSI"]=100-(100/(1+g.rolling(14).mean()/ls.rolling(14).mean().replace(0,np.nan)))
    d30["vm5"]=d30["v"].rolling(5).mean();d30["vr"]=d30["v"]/d30["vm5"].replace(0,np.nan)
    return d30

STOCKS=[
    ("3711","日月光","權值股","盤整","K81超買高檔注意回檔","權值股沿MA5操作，守今日低點680"),
    ("4958","臻鼎KY","穩健股","盤整","K<D偏空量縮防守570","防守30分K MA20約570等量增金叉"),
    ("3042","晶技","績優股","高檔量縮","K值急升量價背離","尾盤強拉無量，嚴守紅K低點"),
    ("2337","旺宏","循環股","低檔","超賣K=17等金叉反彈","記憶體景氣循環，等低檔量增金叉"),
    ("2436","偉詮電","題材股","高檔","高檔盤整守MA20","高檔防守不破續抱"),
    ("3673","TPKKY","活潑股","高檔死叉","買氣93%但高檔死叉","高檔死叉注意防守75元"),
    ("5351","鈺創","主力股","高檔量縮","K=93極高檔量縮","高檔量縮防殺多出清"),
    ("2317","鴻海","權值股","超賣","超賣K=15不殺低","死叉常為多頭假摔，等金叉低接"),
    ("2454","聯發科","龍頭股","回檔","短線死叉日K多頭不變","日K多頭沒破，續抱不被嚇跑"),
    ("8150","南茂","循環股","區間","區間90~120等95以下","等跌到95-98且帶量金叉再進"),
    ("2330","台積電","龍頭股","築底","爆量2.6倍中位盤整","爆量低點不破可拉回切入"),
    ("4961","天鈺","活潑股","蓄勢","爆量1.89倍守165","開盤守165則換手成功噴出"),
    ("6451","訊芯KY","波動股","回檔","等回測500~510","高波動CPO洗盤兇，等止跌金叉"),
    ("3532","台勝科","循環股","起漲🟢","金叉底部成形","底部成形可信度高可試單"),
    ("2327","國巨","循環股","超賣","K=16超賣等金叉","超賣等大量金叉右側反彈"),
    ("8016","矽創","績優股","築底","量能回升大戶吸籌","量能回升大戶暗中吸籌"),
    ("2464","盟立","波動股","超賣","K=13極超賣左側埋伏","極小資金埋伏博暴利反彈"),
    ("6139","亞翔","波動股","高檔🔴","高檔死叉避開","高檔死叉動能轉弱切勿追高"),
    ("1303","南亞","循環股","高檔🔴","高檔死叉量縮避開","高檔量縮死叉波段見頂避開"),
    ("5425","台半","循環股","高檔🔴","K=96量縮假突破","K=96量窒息假突破明日砸盤"),
]

def calc(sid,d30):
    if d30 is None or len(d30)<14: return None
    l=d30.iloc[-1];p=d30.iloc[-2];p2=d30.iloc[-3]
    k=round(l["K"],1);d=round(l["D"],1);kp=round(p["K"],1);dp=round(p["D"],1)
    pr=round(l["c"],2);vr=round(l["vr"],2) if not pd.isna(l.get("vr",np.nan)) else 0
    gap=round(l["KDgap"],1) if not pd.isna(l.get("KDgap",np.nan)) else 0
    gp=round(p["KDgap"],1)
    mh=round(l["MACDh"],2);mhp=round(p["MACDh"],2)
    rsi_v=round(l["RSI"],1) if not pd.isna(l.get("RSI",np.nan)) else 50
    rsi_p=round(p["RSI"],1)
    
    if kp<=dp and k>d:
        if vr>1.0: kd_s="🟢金叉帶量"
        else: kd_s="🟡金叉量縮"
    elif kp>=dp and k<d:
        kd_s="🔴死叉"
    elif gap<0 and gap>gp and gap>-3 and k<50:
        kd_s=f"💡即將金叉(乖離{gap})"
    else:
        kd_s="⚪盤整"
    if k>80: kd_s+="高檔"
    elif k<20: kd_s+="低檔"
    
    if rsi_v>70: rsi_s=f"RSI{rsi_v}超買"
    elif rsi_v<30: rsi_s=f"RSI{rsi_v}超賣"
    elif rsi_v>50: rsi_s=f"RSI{rsi_v}偏多"
    else: rsi_s=f"RSI{rsi_v}偏空"
    
    macdv=round(l["MACD"],2)
    if macdv>0:
        if mh>mhp: macd_s="🔴紅柱↑"
        else: macd_s="🔴紅柱↓"
    else:
        if mh<mhp: macd_s="🟢綠柱↑"
        else: macd_s="🟢綠柱↓"
    if macdv>0 and mh<0: macd_s="🔄剛翻紅"
    elif macdv<0 and mh>0: macd_s="🔄剛翻綠"
    
    hints=[]
    if "金叉帶量" in kd_s: hints.append("🟢金叉+量增+MACD紅=強力買點")
    elif "金叉量縮" in kd_s: hints.append("🟡金叉量縮=假突破90%")
    elif "死叉" in kd_s and vr>1.0: hints.append("🔴死叉放量=建議出場")
    elif "死叉" in kd_s: hints.append("🟠死叉觀察")
    if "即將金叉" in kd_s: hints.append("💡準備埋伏！資金準備好")
    if rsi_v>70 and rsi_v<rsi_p: hints.append("⚠️RSI背離動能減")
    if rsi_v<30 and rsi_v>rsi_p: hints.append("💡RSI背離動能增")
    if k>80 and vr<0.8: hints.append("⚠️高檔量縮主力調節")
    if k<20 and vr>1.0: hints.append("💡超賣放量買盤進")
    if not hints: hints.append("⚪盤整觀望")
    hint=" | ".join(hints[:3])
    
    vs="爆量🔥" if vr>1.5 else ("正常" if vr>0.8 else "量縮")
    
    return {"price":pr,"kd_s":kd_s,"rsi_s":rsi_s,"macd_s":macd_s,"vol_s":vs,"hint":hint}

print("📊產出晨報...")
api=login();rows=[]
for sid,name,ntype,trend,note,strategy in STOCKS:
    d30=get30(api,sid);info=calc(sid,d30)
    if info:
        rows.append([sid,name,ntype,trend]+[info["price"],info["kd_s"],info["rsi_s"],info["macd_s"],info["vol_s"],info["hint"],strategy])
        print(f"  {name} {info['kd_s']} {info['rsi_s']} {info['macd_s']}")
api.logout()

wb=openpyxl.Workbook();ws=wb.active;ws.title=f"{now.strftime('%m%d')}晨報"
hf=Font(bold=True,color="FFFFFF",size=10);hfill=PatternFill(start_color="001A237E",end_color="00283593",fill_type="solid")
gfill=PatternFill(start_color="00E8F5E9",end_color="00E8F5E9",fill_type="solid")
rfill=PatternFill(start_color="00FFEBEE",end_color="00FFEBEE",fill_type="solid")
tb=Border(left=Side(style='thin',color='00DDDDDD'),right=Side(style='thin',color='00DDDDDD'),top=Side(style='thin',color='00DDDDDD'),bottom=Side(style='thin',color='00DDDDDD'))

ws.merge_cells('A1:K1')
ws['A1'].value=f"📊 晨報 {now.strftime('%Y/%m/%d')} — 懂骨性實戰版"
ws['A1'].font=Font(bold=True,size=14,color="001A237E");ws['A1'].alignment=Alignment(horizontal='center')

headers=["代號","名稱","股性","趨勢","股價","30分KD","RSI","MACD","量能","🎯提示","策略"]
for c,h in enumerate(headers,1):
    cell=ws.cell(row=3,column=c,value=h);cell.font=hf;cell.fill=hfill;cell.alignment=Alignment(horizontal='center',wrap_text=True);cell.border=tb

for i,rd in enumerate(rows,4):
    for c,v in enumerate(rd,1):
        cell=ws.cell(row=i,column=c,value=v);cell.border=tb;cell.alignment=Alignment(vertical='center',wrap_text=True)
        if "🔴" in str(rd[5]): cell.fill=rfill
        elif "🟢" in str(rd[5]) or "💡" in str(rd[5]): cell.fill=gfill
    ws.cell(row=i,column=1).font=Font(bold=True);ws.cell(row=i,column=2).font=Font(bold=True)
    ws.cell(row=i,column=10).font=Font(size=10,color="00C62828")
    ws.cell(row=i,column=11).font=Font(size=10,color="001565C0")
    ws.row_dimensions[i].height=30

for c,w in zip(range(1,12),[7,10,8,8,8,22,18,16,8,45,35]):
    ws.column_dimensions[chr(64+c)].width=w

r=len(rows)+5
ws.cell(row=r,column=1,value="📌心法：").font=Font(bold=True,size=10)
for i,t in enumerate([
    "權值股(台積鴻海)：看日K方向，30分K金叉加碼、死叉不殺低",
    "IC設計/高波動(聯發科鈺創)：RSI背離或MACD翻綠立刻走",
    "景氣循環(旺宏南茂)：低檔量增+MACD轉正+金叉才進",
    "績優股(晶技矽創)：沿著MA20操作，跌破就賣",
    "活潑題材股(天鈺盟立)：追量不追價，爆量拉回再接",
    "主力股(鈺創)：K>80+量縮=準備出貨，MACD翻綠就要跑",
]):
    ws.cell(row=r+1+i,column=1,value=t)

wb.save(OUTPUT)
print(f"✅ {OUTPUT}")
